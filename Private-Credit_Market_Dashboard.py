import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from fredapi import Fred

# 🔷 FRED API
fred = Fred(api_key="823ae1695c5aa874894f0973bae91673 ")

# 🔷 Tickers
tickers = {
    "HYG": "HYG",
    "Loan ETF": "SRLN",
    "Blackstone": "BX",
    "Apollo": "APO",
    "Ares": "ARES",
    "KKR": "KKR"
}

# 🔷 Date range
end_date = datetime.today()
start_date = end_date - timedelta(days=14)

data = []

# 🔷 Fetch market data
for name, ticker in tickers.items():
    df = yf.download(ticker, start=start_date, end=end_date)

    if df.empty:
        continue
# handle multi index properly
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)  # Flatten multi-index columns

    df = df.reset_index()
    df = df.sort_values("Date")

# using last 2 rows (current and last week)
    if len(df) >= 2:
        current = float(df.iloc[-1]["Close"])
        last_week = float(df.iloc[-2]["Close"])
        change = (current - last_week) / last_week

        data.append({
            "Metric": name,
            "Current": round(current, 2),
            "Last Week": round(last_week, 2),
            "% Change": round(change * 100, 2)
        })

# 🔷 HY Spread (FRED)
hy_spread = fred.get_series("BAMLH0A0HYM2")

hy_spread = hy_spread.dropna()
hy_spread.index = pd.to_datetime(hy_spread.index)

fridays = hy_spread[hy_spread.index.day_name() == "Friday"]

if len(fridays) >= 2:
    current_spread = float(fridays.iloc[-1])
    last_week_spread = float(fridays.iloc[-2])

    spread_change = (current_spread - last_week_spread) / last_week_spread

    data.append({
        "Metric": "HY Spread",
        "Current": round(current_spread, 2),
        "Last Week": round(last_week_spread, 2),
        "% Change": round(spread_change * 100, 2)
    })

# 🔷 Create DataFrame
output = pd.DataFrame(data)

# 🔷 Dashboard path (make sure this is correct)
dashboard_path = "/Users/nikhilreddy/Desktop/Excel/Private_Credit_Dashboard.xlsx"

# 🔷 Load workbook
book = load_workbook(dashboard_path)
sheet = book["Data_Input"]

# 🔷 Clear old data
for row in sheet.iter_rows(min_row=2, max_row=100, max_col=4):
    for cell in row:
        cell.value = None

# 🔷 Write new data
for i, row in output.iterrows():
    sheet.cell(row=i+2, column=1, value=row["Metric"])
    sheet.cell(row=i+2, column=2, value=row["Current"])
    sheet.cell(row=i+2, column=3, value=row["Last Week"])
    sheet.cell(row=i+2, column=4, value=row["% Change"])

# 🔷 Save
book.save(dashboard_path)

print("Dashboard updated successfully!")